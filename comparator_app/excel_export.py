from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from .models import ComparedRow


LIGHT_RED_FILL = PatternFill(fill_type="solid", fgColor="FFF4CCCC")


def write_output_excel(
    compared_rows: list[ComparedRow],
    model_path: Path,
    output_path: Path,
) -> None:
    workbook = load_workbook(str(model_path))
    sheet = workbook[workbook.sheetnames[0]]

    max_row = sheet.max_row
    for row_index in range(2, max_row + 1):
        for col in range(1, 10):
            cell = sheet.cell(row=row_index, column=col)
            cell.value = None
            cell.fill = PatternFill(fill_type=None)

    field_to_excel_col = {
        "nominal_value": 3,
        "measured_value": 4,
        "lower_limit": 5,
        "upper_limit": 6,
        "deviation": 7,
        "exceedance": 8,
    }

    for i, compared in enumerate(compared_rows, start=2):
        if i > sheet.max_row:
            sheet.append([None] * 9)

        row = compared.row
        sheet.cell(i, 1).value = i - 1
        sheet.cell(i, 2).value = row.characteristic_name
        sheet.cell(i, 3).value = row.nominal_value
        sheet.cell(i, 4).value = row.measured_value
        sheet.cell(i, 5).value = row.lower_limit
        sheet.cell(i, 6).value = row.upper_limit
        sheet.cell(i, 7).value = row.deviation
        sheet.cell(i, 8).value = row.exceedance
        sheet.cell(i, 9).value = compared.status

        for field_name in compared.mismatched_fields:
            excel_col = field_to_excel_col[field_name]
            sheet.cell(i, excel_col).fill = LIGHT_RED_FILL

        if compared.status == "not ok":
            sheet.cell(i, 9).fill = LIGHT_RED_FILL

    workbook.save(str(output_path))
